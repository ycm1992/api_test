# --*-- coding=utf-8 --*--
# @File    : do_excel.py
# @Time    : 2019-05-04 19:07
# @Author  : chuanman.yu
# @Email   : 1165358716@qq.com
# @Software: PyCharm


from openpyxl import load_workbook
from API_2.common.project_path import conf_path
from API_2.common.read_config import ReadConfig
from API_2.common.project_path import case_path


class DoExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_data(self, section):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        case_id = ReadConfig(conf_path).get_origin_data(section, "case_id")
        tel = self.get_tel()
        print("tel的数据类型是：", type(tel))
        test_data = []
        for i in range(2, sheet.max_row+1):
            row_data = {}
            row_data["CaseId"] = sheet.cell(i, 1).value
            row_data["Module"] = sheet.cell(i, 2).value
            row_data["Title"] = sheet.cell(i, 3).value
            row_data["Url"] = sheet.cell(i, 4).value
            row_data["Method"] = sheet.cell(i, 5).value
            # ！=-1，find内置函数在没有查找到结果时会返回-1，所以！=-1是有找到
            if sheet.cell(i, 6).value.find("tel") != -1:
                # 因为从excel中获取的tel是int，所以要转成str才能替换
                row_data["Params"] = sheet.cell(i, 6).value.replace("tel", str(tel))
                self.update_tel(tel+1)
            else:
                row_data["Params"] = sheet.cell(i, 6).value
            row_data["ExpectedResult"] = sheet.cell(i, 8).value
            test_data.append(row_data)
        wb.close()
        final_data = []
        if case_id == "all":
            final_data = test_data
        else:
            for i in case_id:
                final_data.append(test_data[i-1])
        return final_data

    def get_tel(self):
        # 获取手机号码
        wb = load_workbook(self.file_name)
        sheet = wb["tel"]
        wb.close()
        return sheet.cell(1, 2).value

    def update_tel(self, new_tel):
        # 更新手机号码
        wb = load_workbook(self.file_name)
        sheet = wb["tel"]
        sheet.cell(1, 2, new_tel)
        wb.save(self.file_name)
        wb.close()

    def write_back(self, row, col, val):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, col).value = val
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':

    result = DoExcel(case_path, "recharge").read_data("RECHARGE_CASE")
    # result = DoExcel(case_path, "tel").get_tel()
    print(result)






